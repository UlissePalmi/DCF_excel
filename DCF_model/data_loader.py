"""
Data loader module for the YPF DCF Model.
Loads historical and projected data from a CSV/Excel source file.
"""

import csv
import os
from typing import Optional

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DataLoader:
    """
    Loads numeric data from a CSV or Excel file exported from the YPF Model sheet.
    
    The data is stored as a dict-of-dicts keyed by (row, col) 1-indexed coordinates,
    mirroring the Excel layout. Column H=8 corresponds to 2020, I=9 to 2021, etc.
    """

    # Column-to-year mapping (H=8 -> 2020, ..., V=22 -> 2034)
    COL_TO_YEAR = {8: 2020, 9: 2021, 10: 2022, 11: 2023, 12: 2024,
                   13: 2025, 14: 2026, 15: 2027, 16: 2028, 17: 2029,
                   18: 2030, 19: 2031, 20: 2032, 21: 2033, 22: 2034}
    YEAR_TO_COL = {v: k for k, v in COL_TO_YEAR.items()}

    HISTORICAL_YEARS = [2020, 2021, 2022, 2023, 2024]
    PROJECTED_YEARS = [2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034]
    ALL_YEARS = HISTORICAL_YEARS + PROJECTED_YEARS

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._data: dict[tuple[int, int], float | str | None] = {}
        self._load()

    def _load(self):
        ext = os.path.splitext(self.filepath)[1].lower()
        if ext == '.csv':
            self._load_csv()
        elif ext in ('.xlsx', '.xls'):
            self._load_excel()
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def _load_csv(self):
        with open(self.filepath, 'r') as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, start=1):
                for c_idx, val in enumerate(row, start=1):
                    if val:
                        try:
                            self._data[(r_idx, c_idx)] = float(val)
                        except ValueError:
                            self._data[(r_idx, c_idx)] = val

    def _load_excel(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to read Excel files")
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb['Model']
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                max_col=ws.max_column, values_only=False):
            for cell in row:
                if cell.value is not None:
                    self._data[(cell.row, cell.column)] = cell.value

    def get(self, row: int, col: int, default=None):
        """Get a cell value by 1-indexed (row, col)."""
        return self._data.get((row, col), default)

    def get_by_year(self, row: int, year: int, default=None):
        """Get a cell value by row number and year."""
        col = self.YEAR_TO_COL.get(year)
        if col is None:
            return default
        return self.get(row, col, default)

    def get_row_series(self, row: int, years: Optional[list[int]] = None) -> dict[int, float]:
        """Return {year: value} for a given row across specified years."""
        years = years or self.ALL_YEARS
        result = {}
        for y in years:
            v = self.get_by_year(row, y)
            if v is not None:
                result[y] = v
        return result

    def get_historical(self, row: int) -> dict[int, float]:
        return self.get_row_series(row, self.HISTORICAL_YEARS)

    def get_projected(self, row: int) -> dict[int, float]:
        return self.get_row_series(row, self.PROJECTED_YEARS)
