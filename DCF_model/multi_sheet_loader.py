"""
Multi-sheet data loader for the YPF DCF Model.

Reads an Excel workbook where each worksheet corresponds to one schedule.

Sheet format:
    Row 1:  [ignored] | 2020 | 2021 | ... | 2034   (year header)
    Row 2+: field_key | val  | val  | ... | val
"""

import os

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


HISTORICAL_YEARS = [2020, 2021, 2022, 2023, 2024]
PROJECTED_YEARS  = [2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034]
ALL_YEARS        = HISTORICAL_YEARS + PROJECTED_YEARS


class MultiSheetLoader:
    """
    Reads a multi-sheet Excel workbook.  Each sheet holds one schedule's data
    as a flat table: col A = field key, row 1 = year headers, cells = values.
    """

    HISTORICAL_YEARS = HISTORICAL_YEARS
    PROJECTED_YEARS  = PROJECTED_YEARS
    ALL_YEARS        = ALL_YEARS

    def __init__(self, filepath: str):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to read multi-sheet Excel files")
        self.filepath = filepath
        # {sheet_name: {field_key: {year: float}}}
        self._sheets: dict[str, dict[str, dict[int, float]]] = {}
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._sheets[sheet_name] = self._parse_sheet(ws)

    def _parse_sheet(self, ws) -> dict[str, dict[int, float]]:
        """Parse one sheet → {field_key: {year: value}}."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        # Row 0: year header — col 0 is the label column, col 1+ are years
        year_row = rows[0]
        years = []
        for v in year_row[1:]:
            if isinstance(v, (int, float)) and 2000 <= int(v) <= 2100:
                years.append(int(v))

        data = {}
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            values: dict[int, float] = {}
            for i, year in enumerate(years):
                raw = row[i + 1] if i + 1 < len(row) else None
                if raw is not None:
                    try:
                        values[year] = float(raw)
                    except (TypeError, ValueError):
                        pass
            data[key] = values

        return data

    def field(self, sheet_name: str, key: str) -> dict[int, float]:
        """Return {year: value} for the given sheet + field key."""
        sheet = self._sheets.get(sheet_name)
        if sheet is None:
            raise KeyError(
                f"Sheet '{sheet_name}' not found. "
                f"Available: {list(self._sheets)}"
            )
        series = sheet.get(key)
        if series is None:
            raise KeyError(
                f"Field '{key}' not found in sheet '{sheet_name}'. "
                f"Available: {list(sheet)}"
            )
        return series

    @property
    def sheet_names(self) -> list[str]:
        return list(self._sheets.keys())
