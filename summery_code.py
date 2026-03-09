"""
Replicates the 'Summary' sheet from the YPF_DCF.xlsx workbook.
Output: YPF_DCF_Summary.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# ── STYLE DEFINITIONS ─────────────────────────────────────────────────────────
NAVY = '002F6C'

font_title      = Font(name='Arial', size=18, bold=True)
font_subtitle   = Font(name='Arial', size=14, bold=True)
font_section    = Font(name='Arial', size=10, bold=True)
font_hdr_white  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_label      = Font(name='Arial', size=10)
font_sub_label  = Font(name='Arial', size=9)
font_sub_lbl8   = Font(name='Arial', size=8)
font_navy       = Font(name='Arial', size=10, color=NAVY)
font_navy9      = Font(name='Arial', size=9, color=NAVY)
font_navy_bold  = Font(name='Arial', size=10, bold=True, color=NAVY)
font_impl_white = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_impl_dark  = Font(name='Arial', size=10, bold=True, color='404040')
font_pv_bold    = Font(name='Calibri', size=11, bold=True, color='000000')

fill_blue   = PatternFill('solid', fgColor=NAVY)
fill_navy   = PatternFill('solid', fgColor=NAVY)
fill_yellow = PatternFill('solid', fgColor=NAVY)

align_cc    = Alignment(horizontal='centerContinuous')
align_cc_vc = Alignment(horizontal='centerContinuous', vertical='center')
align_vc    = Alignment(vertical='center')

thin_tb = Border(top=Side(style='thin'), bottom=Side(style='thin'))

FMT_DOLLAR  = '"$"#,##0_);\\("$"#,##0\\)'
FMT_DOLLAR2 = '"$"#,##0.00_);\\("$"#,##0.00\\)'
FMT_PCT     = '0.0%;\\(0.0%\\)'
FMT_PCT_S   = '0.0%'
FMT_PCT_W   = '0%'
FMT_YEAR    = '0\\A'
FMT_DEC1    = '#,##0.0_);\\(#,##0.0\\)'
FMT_DEC1S   = '#,##0.0_);\\(#,##0.0\\)'
FMT_DEC1P   = '#,##0.0'
FMT_DEC2    = '#,##0.00_);\\(#,##0.00\\)'
FMT_DEC2S   = '#,##0.00_);(#,##0.00)'
FMT_SHARES  = '#,##0.0'


def sc(ref, value, font=None, fill=None, alignment=None, number_format=None, border=None):
    cell = ws[ref]
    cell.value = value
    if font:          cell.font = font
    if fill:          cell.fill = fill
    if alignment:     cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border:        cell.border = border


# Year columns (left mirror and right source)
LEFT_COLS  = ['H','I','J','K','L','M','N','O','P','Q','R','S','T']
RIGHT_COLS = ['AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT']
PROJ_L     = ['K','L','M','N','O','P','Q','R','S','T']
PROJ_R     = ['AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT']


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1: BASE CASE LEFT PANEL (mirror of right panel)  rows 1-41
# ═══════════════════════════════════════════════════════════════════════════════

sc('B1', '=Cover!B4',    font=font_title,    alignment=align_cc)
sc('B2', 'Base Case DCF', font=font_subtitle, alignment=align_cc)
sc('C5', 'SUMMARY VALUES - BASE CASE', font=font_section, alignment=align_cc)

sc('K6', 'Projected', font=font_hdr_white, fill=fill_blue,
   alignment=align_cc_vc, border=thin_tb)

sc('D7', '($ Millions)', font=font_hdr_white, fill=fill_blue, alignment=align_vc)
sc('F7', 'Trend',        font=font_hdr_white, fill=fill_blue, alignment=align_vc)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}7', f'=${rc}$7', font=font_hdr_white, fill=fill_blue,
       alignment=align_vc, number_format=FMT_YEAR)

sc('D10', 'Income Statement Items', font=font_section)
sc('V10', 'Discount Rate',           font=font_navy)
sc('X10', '=$AX$10',                 font=font_navy, number_format=FMT_PCT_S)

sc('E12', 'Net Revenue', font=font_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}12', f'=${rc}$12', font=font_navy, number_format=FMT_DOLLAR)
sc('V12', 'Terminal Growth Rate', font=font_navy)
sc('X12', '=$AX$12',              font=font_navy, number_format=FMT_PCT_S)

sc('E13', '   Growth', font=font_sub_label)
for col, rc in zip(LEFT_COLS[1:], RIGHT_COLS[1:]):
    sc(f'{col}13', f'=${rc}$13', font=font_navy9, number_format=FMT_PCT)
sc('V13', 'Terminal Value', font=font_navy)
sc('X13', '=$AX$13',        font=font_navy, number_format=FMT_DOLLAR)

sc('E16', 'EBITDA', font=font_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}16', f'=${rc}$16', font=font_navy, number_format=FMT_DOLLAR)
sc('V16', 'Cumulative PV of FCF', font=font_navy)
sc('X16', '=$AX$16',              font=font_navy, number_format=FMT_DOLLAR)

sc('E17', '   Margin', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}17', f'=${rc}$17', font=font_navy9, number_format=FMT_PCT)

sc('E18', '   Growth', font=font_sub_label)
for col, rc in zip(LEFT_COLS[1:], RIGHT_COLS[1:]):
    sc(f'{col}18', f'=${rc}$18', font=font_navy9, number_format=FMT_PCT)

sc('E21', 'Net Income', font=font_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}21', f'=${rc}$21', font=font_navy, number_format=FMT_DOLLAR)
sc('V21', 'PV of Terminal Value', font=font_navy)
sc('X21', '=$AX$21',              font=font_navy, number_format=FMT_DOLLAR)

sc('E22', '   Margin', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}22', f'=${rc}$22', font=font_navy9, number_format=FMT_PCT)

sc('E23', '   Growth', font=font_sub_label)
for col, rc in zip(LEFT_COLS[1:], RIGHT_COLS[1:]):
    sc(f'{col}23', f'=${rc}$23', font=font_navy9, number_format=FMT_PCT)

sc('E26', 'NOPAT', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}26', f'=${rc}$26', font=font_navy, number_format=FMT_DOLLAR)
sc('V26', 'Enterprice Value', font=font_navy)
sc('X26', '=$AX$26',          font=font_navy, number_format=FMT_DOLLAR)

sc('E27', '   Margin', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}27', f'=${rc}$27', font=font_navy9, number_format=FMT_PCT)
sc('V27', 'Net Cash', font=font_navy)
sc('X27', '=$AX$27', font=font_navy, number_format=FMT_DOLLAR)

sc('E28', '   Growth', font=font_sub_label)
for col, rc in zip(LEFT_COLS[1:], RIGHT_COLS[1:]):
    sc(f'{col}28', f'=${rc}$28', font=font_navy9, number_format=FMT_PCT)
sc('V28', 'Equity Value', font=font_navy)
sc('X28', '=$AX$28', font=font_navy, number_format=FMT_DOLLAR)

sc('E31', 'D&A', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}31', f'=${rc}$31', font=font_navy, number_format=FMT_DOLLAR)

sc('E32', 'Capex', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}32', f'=${rc}$32', font=font_navy, number_format=FMT_DOLLAR)
sc('V32', 'Shares (MM)', font=font_navy)
sc('X32', '=$AX$32', font=font_navy, number_format=FMT_SHARES)

sc('E33', 'NWC', font=font_sub_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}33', f'=${rc}$33', font=font_navy, number_format=FMT_DOLLAR)

sc('E36', 'Unlevered FCFF', font=font_label)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}36', f'=${rc}$36', font=font_navy, number_format=FMT_DOLLAR)
sc('V36', 'Implied Shared Price', font=font_impl_white, fill=fill_blue)
sc('X36', '=$AX$36', font=font_impl_white, fill=fill_blue, number_format=FMT_DOLLAR2)

sc('E39', '   Discount Period', font=font_label)
for col, rc in zip(PROJ_L, PROJ_R):
    sc(f'{col}39', f'=${rc}$39', font=font_navy, number_format=FMT_DEC1)

sc('E40', '   Discount Factor', font=font_label)
for col, rc in zip(PROJ_L, PROJ_R):
    sc(f'{col}40', f'=${rc}$40', font=font_navy, number_format=FMT_DEC2)

sc('E41', 'Present Value of FCF', font=font_pv_bold)
for col, rc in zip(PROJ_L, PROJ_R):
    sc(f'{col}41', f'=${rc}$41', font=font_navy, number_format=FMT_DOLLAR)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1B: BASE CASE RIGHT PANEL (source data)  rows 5-41, cols AC-AX
# ═══════════════════════════════════════════════════════════════════════════════

sc('AC5', 'CURRENT CASE: Base Case Scenario', font=font_section, alignment=align_cc)
sc('AK6', 'Projected', font=font_section, alignment=align_cc_vc)

sc('AD10', 'Income Statement Items', font=font_section)
sc('AV10', 'Discount Rate',           font=font_label)
sc('AX10', 0.138,                      font=font_label, number_format=FMT_PCT_S)

# Year headers row 7
right_yr_formulas = [
    '=AI7-1','=AJ7-1','=AK7-1','=Assumtions!G8',
    '=AK7+1','=AL7+1','=AM7+1','=AN7+1','=AO7+1','=AP7+1','=AQ7+1','=AR7+1','=AS7+1'
]
for col, formula in zip(RIGHT_COLS, right_yr_formulas):
    sc(f'{col}7', formula, font=font_section, alignment=align_vc, number_format=FMT_YEAR)

# Net Revenue
sc('AE12', 'Net Revenue', font=font_label)
sc('AF12', '(MM)',         font=font_sub_lbl8)
model_cols = ['J','K','L','M','N','O','P','Q','R','S','T','U','V']
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}12', f'=Model!{mc}401', font=font_label, number_format=FMT_DOLLAR)
sc('AV12', 'Terminal Growth Rate', font=font_label)
sc('AX12', 0.035,                      font=font_label, number_format=FMT_PCT_S)

sc('AE13', '   Growth', font=font_sub_label)
sc('AF13', '(%)',        font=font_sub_label)
for i, col in enumerate(RIGHT_COLS[1:], 1):
    prev = RIGHT_COLS[i-1]
    sc(f'{col}13', f'={col}12/{prev}12-1', font=font_sub_label, number_format=FMT_PCT)
sc('AV13', 'Terminal Value',              font=font_label)
sc('AX13', '=AT36*(1+AX12)/(AX10-AX12)', font=font_label, number_format=FMT_DOLLAR)

# EBITDA
sc('AE16', 'EBITDA', font=font_label)
sc('AF16', '(MM)',    font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}16', f'=Model!{mc}412', font=font_label, number_format=FMT_DOLLAR)
sc('AV16', 'Cumulative PV of FCF', font=font_label)
sc('AX16', '=SUM(AK41:AT41)',      font=font_label, number_format=FMT_DOLLAR)

sc('AE17', '   Margin', font=font_sub_label)
sc('AF17', '(%)',        font=font_sub_label)
for col in RIGHT_COLS:
    sc(f'{col}17', f'={col}16/{col}12', font=font_sub_label, number_format=FMT_PCT)

sc('AE18', '   Growth', font=font_sub_label)
sc('AF18', '(%)',        font=font_sub_label)
for i, col in enumerate(RIGHT_COLS[1:], 1):
    prev = RIGHT_COLS[i-1]
    sc(f'{col}18', f'={col}16/{prev}16-1', font=font_sub_label, number_format=FMT_PCT)

# Net Income
sc('AE21', 'Net Income', font=font_label)
sc('AF21', '(MM)',        font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}21', f'=Model!{mc}425', font=font_label, number_format=FMT_DOLLAR)
sc('AV21', 'PV of Terminal Value', font=font_label)
sc('AX21', '=AX13*AT40',           font=font_label, number_format=FMT_DOLLAR)

sc('AE22', '   Margin', font=font_sub_label)
sc('AF22', '(%)',        font=font_sub_label)
for col in RIGHT_COLS:
    sc(f'{col}22', f'={col}21/{col}12', font=font_sub_label, number_format=FMT_PCT)

sc('AE23', '   Growth', font=font_sub_label)
sc('AF23', '(%)',        font=font_sub_label)
for i, col in enumerate(RIGHT_COLS[1:], 1):
    prev = RIGHT_COLS[i-1]
    sc(f'{col}23', f'={col}21/{prev}21-1', font=font_sub_label, number_format=FMT_PCT)

# NOPAT
sc('AE26', 'NOPAT', font=font_sub_label)
sc('AF26', '(MM)',   font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}26', f'=Model!{mc}431', font=font_label, number_format=FMT_DOLLAR)
sc('AV26', 'Enterprice Value', font=font_label)
sc('AX26', '=AX16+AX21',       font=font_label, number_format=FMT_DOLLAR)

sc('AE27', '   Margin', font=font_sub_label)
sc('AF27', '(%)',        font=font_sub_label)
for col in RIGHT_COLS:
    sc(f'{col}27', f'={col}26/{col}12', font=font_sub_label, number_format=FMT_PCT)
sc('AV27', 'Net Cash',                                     font=font_label)
sc('AX27', '=Model!M502-Model!M538-Model!M525-Model!M551', font=font_label, number_format=FMT_DOLLAR)

sc('AE28', '   Growth', font=font_sub_label)
sc('AF28', '(%)',        font=font_sub_label)
for i, col in enumerate(RIGHT_COLS[1:], 1):
    prev = RIGHT_COLS[i-1]
    sc(f'{col}28', f'={col}26/{prev}26-1', font=font_sub_label, number_format=FMT_PCT)
sc('AV28', 'Equity Value',    font=font_label)
sc('AX28', '=SUM(AX26:AX27)', font=font_label, number_format=FMT_DOLLAR)

# D&A
sc('AE31', 'D&A',  font=font_sub_label)
sc('AF31', '(MM)', font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}31', f'=Model!{mc}451', font=font_label, number_format=FMT_DOLLAR)

# Capex
sc('AE32', 'Capex', font=font_sub_label)
sc('AF32', '(MM)',  font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}32', f'=Model!{mc}466', font=font_label, number_format=FMT_DOLLAR)
sc('AV32', 'Shares (MM)',     font=font_label)
sc('AX32', '=Model!L749/1000', font=font_label, number_format='#,##0_);(#,##0)')

# NWC
sc('AE33', 'NWC',  font=font_sub_label)
sc('AF33', '(MM)', font=font_sub_lbl8)
for col, mc in zip(RIGHT_COLS, model_cols):
    sc(f'{col}33', f'=Model!{mc}462', font=font_label, number_format=FMT_DOLLAR)

# Unlevered FCFF
sc('AE36', 'Unlevered FCFF', font=font_label)
sc('AF36', '(MM)',            font=font_sub_lbl8)
for col in RIGHT_COLS:
    sc(f'{col}36', f'=SUM({col}26,{col}31:{col}33)', font=font_label, number_format=FMT_DOLLAR)
sc('AV36', 'Implied Shared Price', font=font_impl_dark)
sc('AX36', '=MAX(AX28/AX32)',      font=font_impl_dark, number_format=FMT_DOLLAR2)

# Discount Period
sc('AE39', '   Discount Period', font=font_label)
sc('AF39', '(MM)',               font=font_sub_lbl8)
sc('AK39', 0.5, font=font_label, number_format=FMT_DEC1S)
prev_dp = 'AK'
for col in ['AL','AM','AN','AO','AP','AQ','AR','AS','AT']:
    sc(f'{col}39', f'={prev_dp}39+1', font=font_label, number_format=FMT_DEC1S)
    prev_dp = col

# Discount Factor
sc('AE40', '   Discount Factor', font=font_label)
sc('AF40', '(MM)',               font=font_sub_lbl8)
for col in PROJ_R:
    sc(f'{col}40', f'=1/(1+$AX$10)^{col}39', font=font_label, number_format=FMT_DEC2S)

# Present Value of FCF
sc('AE41', 'Present Value of FCF', font=font_pv_bold)
sc('AF41', '(MM)',                  font=font_sub_lbl8)
for col in PROJ_R:
    sc(f'{col}41', f'={col}36*{col}40', font=font_label, number_format=FMT_DOLLAR)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2: BEST CASE  rows 45-85
# ═══════════════════════════════════════════════════════════════════════════════

sc('B45', '=B1',          font=font_title,    alignment=align_cc)
sc('B46', 'Best Case DCF', font=font_subtitle, alignment=align_cc)
sc('C49', 'SUMMARY VALUES - BEST CASE', font=font_section, alignment=align_cc)

sc('K50', 'Projected', font=font_hdr_white, fill=fill_navy,
   alignment=align_cc_vc, border=thin_tb)

sc('D51', '($ Millions)', font=font_hdr_white, fill=fill_navy, alignment=align_vc)
sc('F51', 'Trend',        font=font_hdr_white, fill=fill_navy, alignment=align_vc)
for col, rc in zip(LEFT_COLS, RIGHT_COLS):
    sc(f'{col}51', f'=${rc}$7', font=font_hdr_white, fill=fill_navy,
       alignment=align_vc, number_format=FMT_YEAR)

sc('D54', 'Income Statement Items', font=font_section)
sc('V54', 'Discount Rate',           font=font_navy)

# Labels only — no data values (Best Case data columns are empty in original)
sc('E56', 'Net Revenue',      font=font_label)
sc('V56', 'Terminal Growth Rate', font=font_navy)
sc('E57', '   Growth',        font=font_sub_label)
sc('V57', 'Terminal Value',   font=font_navy)
sc('E60', 'EBITDA',           font=font_label)
sc('V60', 'Cumulative PV of FCF', font=font_navy)
sc('E61', '   Margin',        font=font_sub_label)
sc('E62', '   Growth',        font=font_sub_label)
sc('E65', 'Net Income',       font=font_label)
sc('V65', 'PV of Terminal Value', font=font_navy)
sc('E66', '   Margin',        font=font_sub_label)
sc('E67', '   Growth',        font=font_sub_label)
sc('E70', 'NOPAT',            font=font_sub_label)
sc('V70', 'Enterprice Value', font=font_navy)
sc('E71', '   Margin',        font=font_sub_label)
sc('V71', 'Net Cash',         font=font_navy)
sc('E72', '   Growth',        font=font_sub_label)
sc('V72', 'Equity Value',     font=font_navy)
sc('E75', 'D&A',              font=font_sub_label)
sc('E76', 'Capex',            font=font_sub_label)
sc('V76', 'Shares (MM)',      font=font_navy)
sc('E77', 'NWC',              font=font_sub_label)
sc('E80', 'Unlevered FCFF',   font=font_label)
sc('V80', 'Implied Shared Price', font=Font(name='Arial', size=10, bold=True, color='000000'),
   fill=fill_yellow)
sc('E83', '   Discount Period', font=font_label)
sc('E84', '   Discount Factor', font=font_label)
sc('E85', 'Present Value of FCF', font=font_pv_bold)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3: WORST CASE  rows 89-129
# ═══════════════════════════════════════════════════════════════════════════════

sc('B89', '=B1',           font=font_title,    alignment=align_cc)
sc('B90', 'Worst Case DCF', font=font_subtitle, alignment=align_cc)
sc('C93', 'SUMMARY VALUES - WORST CASE', font=font_section, alignment=align_cc)

sc('K94', 'Projected', font=font_hdr_white, fill=fill_navy, alignment=align_cc_vc)

sc('D95', '($ Millions)', font=font_hdr_white, fill=fill_navy, alignment=align_vc)
sc('F95', 'Trend',        font=font_hdr_white, fill=fill_navy, alignment=align_vc)
# Worst case only shows 8 years (2022-2029), hardcoded as values in original
wc_years = [2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029]
for col, yr in zip(['H','I','J','K','L','M','N','O'], wc_years):
    sc(f'{col}95', yr, font=font_hdr_white, fill=fill_navy, alignment=align_vc)

sc('D98', 'Income Statement Items', font=font_section)
sc('V98', 'Discount Rate',           font=font_label)
sc('X98', 0.1,                       font=font_label, number_format=FMT_PCT_W)

# Labels and sidebar values only — data columns are empty in original
sc('E100', 'Net Revenue',         font=font_label)
sc('V100', 'Terminal Growth Rate', font=font_label)
sc('X100', 0,                     font=font_label, number_format=FMT_PCT_W)
sc('E101', '   Growth',           font=font_sub_label)
sc('V101', 'Terminal Value',      font=font_label)
sc('X101', -372.54239064484847,   font=font_label, number_format=FMT_DOLLAR)

sc('E104', 'EBITDA',              font=font_label)
sc('V104', 'Cumulative PV of FCF', font=font_label)
sc('X104', 3.979418952317271,     font=font_label, number_format=FMT_DOLLAR)
sc('E105', '   Margin',           font=font_sub_label)
sc('E106', '   Growth',           font=font_sub_label)

sc('E109', 'Net Income',          font=font_label)
sc('V109', 'PV of Terminal Value', font=font_label)
sc('X109', -242.60995313706715,   font=font_label, number_format=FMT_DOLLAR)
sc('E110', '   Margin',           font=font_sub_label)
sc('E111', '   Growth',           font=font_sub_label)

sc('E114', 'NOPAT',               font=font_sub_label)
sc('V114', 'Enterprice Value',    font=font_label)
sc('X114', -238.63053418474988,   font=font_label, number_format=FMT_DOLLAR)
sc('E115', '   Margin',           font=font_sub_label)
sc('V115', 'Net Cash',            font=font_label)
sc('X115', -126.99999999999991,   font=font_label, number_format=FMT_DOLLAR)
sc('E116', '   Growth',           font=font_sub_label)
sc('V116', 'Equity Value',        font=font_label)
sc('X116', -365.63053418474976,   font=font_label, number_format=FMT_DOLLAR)

sc('E119', 'D&A',                 font=font_sub_label)
sc('E120', 'Capex',               font=font_sub_label)
sc('V120', 'Shares (MM)',         font=font_label)
sc('X120', 40.32937,              font=font_label, number_format=FMT_DOLLAR)
sc('E121', 'NWC',                 font=font_sub_label)

sc('E124', 'Unlevered FCFF',      font=font_label)
sc('V124', 'Implied Shared Price', font=Font(name='Arial', size=10, bold=True, color='000000'),
   fill=fill_yellow)
sc('X124', 0,                     font=Font(name='Arial', size=10, bold=True, color='000000'),
   fill=fill_yellow, number_format=FMT_DOLLAR2)

sc('E127', '   Discount Period',  font=font_label)
sc('E128', '   Discount Factor',  font=font_label)
sc('E129', 'Present Value of FCF', font=font_pv_bold)


# ── SAVE ──────────────────────────────────────────────────────────────────────
output_path = '/mnt/user-data/outputs/YPF_DCF_Summary.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")